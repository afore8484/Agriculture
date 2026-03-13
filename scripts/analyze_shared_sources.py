#!/usr/bin/env python3
from __future__ import annotations

import csv
import json
import re
from collections import Counter
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from pypdf import PdfReader

BASE_DIR = Path(r"E:\workspaces\Agriculture\prototype")
SOURCE_DIR = Path(r"D:\腾讯电脑管家软件搬家\C盘清理文件搬家\xwechat_files\aforecool_c63c\msg\file\2026-03")
XLSX_PATH = SOURCE_DIR / "三农一期共享目录清单-20260101.xlsx"
PDF_PATH = SOURCE_DIR / "053-XX01-018-001.pdf"
OUT_DIR = BASE_DIR / "output" / "data_collection"
REQ_DIR = BASE_DIR / "require" / "数据采集"
JSON_PATH = OUT_DIR / "shared_source_analysis.json"
CSV_PATH = OUT_DIR / "shared_source_tables.csv"
MD_PATH = REQ_DIR / "三农一期共享数据采集需提取表清单.md"


def norm_table_name(name: str | None) -> str:
    if not name:
        return ""
    text = str(name).strip().lower()
    text = re.sub(r"^o_hn\d+_sjsn_", "", text)
    return text


def extraction_advice(record: dict[str, Any]) -> str:
    table = str(record["目标表"] or "")
    if record["PDF匹配页"]:
        return "可直接纳入首批采集，优先按表全量+增量策略实现"
    if re.match(r"^[0-9a-f]{8}_[0-9a-f_]{20,}$", record["标准化表名"]):
        return "表名疑似系统生成对象，需先到库中核实真实业务含义和主键"
    if table.startswith("V_"):
        return "疑似视图对象，建议先确认其基础表和刷新方式"
    return "PDF 未定位到定义，需先做数据库实查和字段核对"


OUT_DIR.mkdir(parents=True, exist_ok=True)
REQ_DIR.mkdir(parents=True, exist_ok=True)

wb = load_workbook(XLSX_PATH, data_only=True)
ws = wb["Sheet1"]
records: list[dict[str, Any]] = []
for row in ws.iter_rows(min_row=2, values_only=True):
    if not any(row):
        continue
    rec = {
        "目录ID": row[0],
        "信息资源名称": row[1],
        "所属系统名称": row[2],
        "对应处室": row[3],
        "数据使用情况": row[4],
        "目标表": row[5],
        "数据库所在IP地址": row[6],
        "共享属性": row[7],
        "开放属性": row[8],
        "是否已挂接": row[9],
        "数据总量": row[10],
        "今年更新量": row[11],
    }
    rec["标准化表名"] = norm_table_name(rec["目标表"])
    records.append(rec)

reader = PdfReader(str(PDF_PATH))
pages: list[str] = []
for page in reader.pages:
    pages.append((page.extract_text() or "").lower())

for rec in records:
    target = rec["标准化表名"]
    matched_page = None
    if target:
        for idx, text in enumerate(pages, 1):
            if target in text:
                matched_page = idx
                break
    rec["PDF匹配页"] = matched_page
    rec["提取建议"] = extraction_advice(rec)

ip_counter = Counter(rec["数据库所在IP地址"] for rec in records)
share_counter = Counter(rec["共享属性"] for rec in records)
open_counter = Counter(rec["开放属性"] for rec in records)
matched_count = sum(1 for rec in records if rec["PDF匹配页"])
unmatched_count = len(records) - matched_count

integration_table_lines: list[str] = []
for idx, text in enumerate(pages, 1):
    if 5 <= idx <= 10:
        for line in text.splitlines():
            line = line.strip()
            if re.search(r"(BASEDATA|META_|DIR_APP_|INTERFACE_|JOB_|SYS_DICT|SYS_LOG|DATA_CENTER_)", line):
                integration_table_lines.append(f"P{idx}: {line}")

payload = {
    "meta": {
        "xlsx_path": str(XLSX_PATH),
        "pdf_path": str(PDF_PATH),
        "record_count": len(records),
        "pdf_match_count": matched_count,
        "pdf_unmatch_count": unmatched_count,
    },
    "summary": {
        "ip_counter": dict(ip_counter),
        "share_counter": dict(share_counter),
        "open_counter": dict(open_counter),
    },
    "records": records,
    "integration_table_lines": integration_table_lines,
}
JSON_PATH.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")

with CSV_PATH.open("w", newline="", encoding="utf-8-sig") as f:
    writer = csv.DictWriter(
        f,
        fieldnames=[
            "目录ID", "信息资源名称", "所属系统名称", "对应处室", "目标表", "标准化表名",
            "数据库所在IP地址", "共享属性", "开放属性", "是否已挂接", "PDF匹配页", "提取建议"
        ],
    )
    writer.writeheader()
    for rec in records:
        writer.writerow({k: rec.get(k) for k in writer.fieldnames})

lines: list[str] = []
lines.append("# 三农一期共享数据采集需提取表清单")
lines.append("")
lines.append("## 1. 分析结论")
lines.append("")
lines.append(f"- 共享目录清单有效记录数：`{len(records)}`")
lines.append(f"- 已挂接记录数：`{sum(1 for r in records if r['是否已挂接'] == '已挂接')}`")
lines.append(f"- PDF 可定位到表定义的记录数：`{matched_count}`")
lines.append(f"- PDF 暂未定位到表定义的记录数：`{unmatched_count}`")
lines.append("- 结论：采集脚本的主采集范围应以 Excel 清单中的 `目标表` 为准，PDF 用于辅助定位表定义、模块归属和数据整合元数据。")
lines.append("")
lines.append("## 2. 数据源概况")
lines.append("")
lines.append("### 2.1 来源文件")
lines.append("")
lines.append(f"- Excel：`{XLSX_PATH}`")
lines.append(f"- PDF：`{PDF_PATH}`")
lines.append("")
lines.append("### 2.2 数据库来源分布")
lines.append("")
lines.append("| 数据库IP | 目录表数量 |")
lines.append("|---|---:|")
for ip, count in sorted(ip_counter.items(), key=lambda kv: (-kv[1], str(kv[0]))):
    lines.append(f"| {ip} | {count} |")
lines.append("")
lines.append("### 2.3 共享与开放属性")
lines.append("")
lines.append("| 维度 | 分类 | 数量 |")
lines.append("|---|---|---:|")
for key, value in share_counter.items():
    lines.append(f"| 共享属性 | {key} | {value} |")
for key, value in open_counter.items():
    lines.append(f"| 开放属性 | {key} | {value} |")
lines.append("")
lines.append("## 3. 采集脚本范围建议")
lines.append("")
lines.append("### 3.1 首批直接采集表")
lines.append("")
lines.append("- 范围：Excel 清单中的全部 `目标表`。")
lines.append("- 原因：这 174 张表均已在共享目录中标记为 `已挂接`，是已确认对接的共享资源。")
lines.append("- 实施优先级：先按数据库 IP 分批，再按表做全量首采和增量策略设计。")
lines.append("")
lines.append("### 3.2 需要重点核实的对象")
lines.append("")
lines.append("- PDF 未定位的 36 张表，需要先到库里核对是否为视图、临时表、汇总表或命名变更后的对象。")
lines.append("- 目标表以 `V_` 开头的对象，优先判断是否为视图，避免直接把视图当物理源表处理。")
lines.append("- 目标表名为 GUID 样式的对象，需要先确认业务含义、主键和更新时间字段。")
lines.append("")
lines.append("### 3.3 PDF 可作为辅助的管理类表")
lines.append("")
lines.append("- PDF 第 4.7 节列出大量数据整合管理表，如 `BASEDATA*`、`META_*`、`DIR_APP_*`、`INTERFACE_*`、`JOB_*`。")
lines.append("- 这类表更适合用于采集平台配置、元数据、任务日志、血缘关系和接口作业管理，不建议与业务共享表混为同一批采集对象。")
lines.append("")
lines.append("## 4. 需提取表清单")
lines.append("")
lines.append("| 目录ID | 信息资源名称 | 目标表 | 标准化表名 | 数据库IP | 共享属性 | 开放属性 | PDF匹配页 | 提取建议 |")
lines.append("|---:|---|---|---|---|---|---|---:|---|")
for rec in records:
    page = rec["PDF匹配页"] or ""
    dept = rec["对应处室"] or ""
    lines.append(
        f"| {rec['目录ID']} | {rec['信息资源名称']} | {rec['目标表']} | {rec['标准化表名']} | {rec['数据库所在IP地址']} | {rec['共享属性']} | {rec['开放属性']} | {page} | {rec['提取建议']} |"
    )
lines.append("")
lines.append("## 5. PDF 中可辅助脚本设计的管理类表")
lines.append("")
lines.append("以下为 PDF 第 4.7 节中抽取到的典型数据整合/元数据/作业管理表，适合作为后续采集平台辅助脚本的数据源：")
lines.append("")
for item in integration_table_lines[:80]:
    lines.append(f"- {item}")
lines.append("")
lines.append("## 6. 后续脚本设计建议")
lines.append("")
lines.append("1. 先按数据库 IP 生成连接配置，再按表清单生成采集任务。")
lines.append("2. 每张表至少补齐 4 个脚本参数：主键字段、增量字段、全量周期、落库目标。")
lines.append("3. 对 PDF 未定位的表，先跑一次数据库元数据探测，自动抓取字段、索引和注释。")
lines.append("4. 对共享属性为“有条件共享”、开放属性为“不予开放”的表，在脚本层单独打权限标签。")

MD_PATH.write_text("\n".join(lines), encoding="utf-8")
print(MD_PATH)
print(JSON_PATH)
print(CSV_PATH)
